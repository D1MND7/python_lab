import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    
    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path)
    
    
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")
   
    try:
        with csv_path.open('r', encoding='utf-8') as csv_file:
           
            sample = csv_file.read(1024)
            csv_file.seek(0)
            
            sniffer = csv.Sniffer()
            dialect = sniffer.sniff(sample)
            has_header = sniffer.has_header(sample)
            
            if not has_header:
                raise ValueError("CSV файл должен содержать заголовок")
            
            reader = csv.reader(csv_file, dialect=dialect)
            rows = list(reader)
            
    except Exception as e:
        raise ValueError(f"Ошибка чтения CSV: {e}")
    
    
    if not rows:
        raise ValueError("Пустой CSV файл")
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
     
        for row in rows:
            ws.append(row)
        
       
        for col_idx, _ in enumerate(rows[0], 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row_idx, row in enumerate(rows, 1):
                cell_value = str(row[col_idx - 1]) if len(row) >= col_idx else ""
                max_length = max(max_length, len(cell_value))
            
            
            adjusted_width = max(max_length + 2, 8)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        
        wb.save(str(xlsx_path))
        
    except Exception as e:
        raise ValueError(f"Ошибка создания XLSX: {e}")